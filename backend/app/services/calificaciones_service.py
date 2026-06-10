import csv
import io
import uuid
from collections.abc import Sequence

from fastapi import HTTPException, UploadFile
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.audit_codes import CALIFICACIONES_IMPORTAR
from app.repositories.calificacion_repository import CalificacionRepository
from app.repositories.padron_repository import EntradaPadronRepository, VersionPadronRepository
from app.repositories.umbral_repository import UmbralMateriaRepository
from app.schemas.auth import UserContext
from app.schemas.calificaciones import (
    VALIDOS_TEXTUALES_APROBATORIOS,
    ActividadDetectada,
    ActividadAImportar,
    CalificacionConfirmResponse,
    CalificacionPreviewResponse,
    CalificacionResponse,
    EntregaSinCorregir,
    ReporteFinalizacionPreviewResponse,
)
from app.services.audit_service import AuditService

COLUMNAS_IGNORAR = frozenset({'alumno', 'apellido', 'nombre', 'apellidos', 'email', 'comision', 'regional', 'legajo'})


def _detectar_actividades(headers: list[str]) -> list[ActividadDetectada]:
    actividades = []
    for h in headers:
        clean = h.strip()
        if clean.lower() in COLUMNAS_IGNORAR:
            continue
        if clean.endswith('(Real)'):
            nombre = clean[:-7].strip()
            tipo = 'numerica'
        else:
            nombre = clean
            tipo = 'textual'
        actividades.append(ActividadDetectada(
            nombre=nombre,
            tipo=tipo,
        ))
    return actividades


def _parse_calificaciones_xlsx(file_bytes: bytes) -> tuple[list[str], list[list]]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail='El archivo está vacío')
    headers = [str(h).strip() if h else '' for h in rows[0]]
    data_rows = []
    for row in rows[1:]:
        if not any(row):
            continue
        data_rows.append(list(row))
    return headers, data_rows


def _parse_calificaciones_csv(file_bytes: bytes) -> tuple[list[str], list[list]]:
    text = file_bytes.decode('utf-8-sig')
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        raise HTTPException(status_code=400, detail='El archivo está vacío')
    headers = [h.strip() for h in rows[0]]
    data_rows = []
    for row in rows[1:]:
        if not any(row):
            continue
        data_rows.append(row)
    return headers, data_rows


def _get_valor_columna(row: list, idx: int) -> str:
    if idx < len(row) and row[idx] is not None:
        return str(row[idx]).strip()
    return ''


def _calcular_aprobado(
    nota_numerica: float | None,
    nota_textual: str | None,
    umbral_pct: int,
    valores_aprobatorios: set[str] | None,
) -> bool | None:
    if nota_numerica is not None:
        return float(nota_numerica) >= umbral_pct
    if nota_textual:
        aprobatorios = valores_aprobatorios or VALIDOS_TEXTUALES_APROBATORIOS
        return nota_textual in aprobatorios
    return None


class CalificacionesService:
    def __init__(self, session: AsyncSession, tenant_id: uuid.UUID):
        self._session = session
        self._tenant_id = tenant_id
        self._calificacion_repo = CalificacionRepository(session, tenant_id)
        self._version_repo = VersionPadronRepository(session, tenant_id)
        self._entrada_repo = EntradaPadronRepository(session, tenant_id)
        self._umbral_repo = UmbralMateriaRepository(session, tenant_id)

    async def preview_import(
        self,
        materia_id: uuid.UUID,
        cohorte_id: uuid.UUID,
        archivo: UploadFile,
    ) -> CalificacionPreviewResponse:
        content = await archivo.read()
        filename = (archivo.filename or '').lower()

        if filename.endswith('.xlsx'):
            headers, data_rows = _parse_calificaciones_xlsx(content)
        elif filename.endswith('.csv'):
            headers, data_rows = _parse_calificaciones_csv(content)
        else:
            raise HTTPException(
                status_code=400,
                detail='Formato no soportado. Usar archivos .xlsx o .csv',
            )

        actividades = _detectar_actividades(headers)

        activos = [a for a in actividades if a.nombre]
        if not activos:
            raise HTTPException(
                status_code=400,
                detail='No se detectaron columnas de actividades en el archivo',
            )

        # Build sample alumno rows for preview
        id_cols = [i for i, h in enumerate(headers) if h.lower() in COLUMNAS_IGNORAR]
        alumnos_ejemplo = []
        for data_row in data_rows[:5]:
            ejemplo = {}
            for idx in id_cols:
                ejemplo[headers[idx]] = _get_valor_columna(data_row, idx)
            for act in activos[:3]:
                col_idx = _find_header_index(headers, act.nombre)
                if col_idx >= 0:
                    ejemplo[act.nombre] = _get_valor_columna(data_row, col_idx)
            alumnos_ejemplo.append(ejemplo)

        return CalificacionPreviewResponse(
            materia_id=materia_id,
            cohorte_id=cohorte_id,
            total_filas=len(data_rows),
            actividades=activos,
            alumnos_ejemplo=alumnos_ejemplo,
        )

    async def confirm_import(
        self,
        request: 'CalificacionConfirmRequest',
        archivo: UploadFile,
        current_user: UserContext,
        audit: AuditService,
    ) -> CalificacionConfirmResponse:
        content = await archivo.read()
        filename = (archivo.filename or '').lower()

        if filename.endswith('.xlsx'):
            headers, data_rows = _parse_calificaciones_xlsx(content)
        elif filename.endswith('.csv'):
            headers, data_rows = _parse_calificaciones_csv(content)
        else:
            raise HTTPException(
                status_code=400,
                detail='Formato no soportado. Usar archivos .xlsx o .csv',
            )

        actividades_detectadas = _detectar_actividades(headers)
        seleccionadas_nombres = {a.nombre for a in request.actividades}
        actividades_a_procesar = [
            a for a in actividades_detectadas if a.nombre in seleccionadas_nombres
        ]

        version_activa = await self._version_repo.get_activa(
            request.materia_id, request.cohorte_id,
        )
        if not version_activa:
            raise HTTPException(
                status_code=400,
                detail='No hay un padrón activo para esta materia y cohorte',
            )

        entradas = await self._entrada_repo.get_by_version(version_activa.id)
        if not entradas:
            raise HTTPException(
                status_code=400,
                detail='El padrón activo no tiene entradas de alumnos',
            )

        entrada_map: dict[str, uuid.UUID] = {}
        for e in entradas:
            key = f'{e.nombre.lower()}|{e.apellidos.lower()}|{e.email.lower()}'
            entrada_map[key] = e.id

        id_cols_lower = {h.lower() for i, h in enumerate(headers) if h.lower() in COLUMNAS_IGNORAR}
        id_cols = {h for h in headers if h.lower() in id_cols_lower}
        col_nombre = _find_header_index(headers, 'nombre')
        if col_nombre < 0:
            col_nombre = _find_header_index(headers, 'Nombre')
        col_apellidos = _find_header_index(headers, 'apellidos')
        if col_apellidos < 0:
            col_apellidos = _find_header_index(headers, 'apellido')
        if col_apellidos < 0:
            col_apellidos = _find_header_index(headers, 'Apellidos')
        col_email = _find_header_index(headers, 'email')
        if col_email < 0:
            col_email = _find_header_index(headers, 'Email')

        calificaciones_data = []
        filas_sin_match = 0

        for data_row in data_rows:
            nombre_val = _get_valor_columna(data_row, col_nombre) if col_nombre >= 0 else ''
            apellidos_val = _get_valor_columna(data_row, col_apellidos) if col_apellidos >= 0 else ''
            email_val = _get_valor_columna(data_row, col_email) if col_email >= 0 else ''
            key = f'{nombre_val.lower()}|{apellidos_val.lower()}|{email_val.lower()}'

            entrada_id = entrada_map.get(key)
            if entrada_id is None:
                filas_sin_match += 1
                continue

            for act in actividades_a_procesar:
                col_idx = _find_header_index_by_actividad(headers, act.nombre)
                if col_idx < 0:
                    continue
                raw_val = _get_valor_columna(data_row, col_idx)
                if not raw_val:
                    continue

                nota_numerica = None
                nota_textual = None
                if act.tipo == 'numerica':
                    try:
                        nota_numerica = float(raw_val.replace(',', '.'))
                    except (ValueError, AttributeError):
                        nota_textual = raw_val
                else:
                    nota_textual = raw_val

                calificaciones_data.append({
                    'entrada_padron_id': entrada_id,
                    'materia_id': request.materia_id,
                    'cohorte_id': request.cohorte_id,
                    'actividad': act.nombre,
                    'nota_numerica': nota_numerica,
                    'nota_textual': nota_textual,
                    'origen': 'Importado',
                })

        if not calificaciones_data:
            raise HTTPException(
                status_code=400,
                detail='No se pudieron vincular calificaciones con el padrón activo',
            )

        await self._calificacion_repo.bulk_create(calificaciones_data)

        await audit.log(
            accion=CALIFICACIONES_IMPORTAR,
            detalle={
                'materia_id': str(request.materia_id),
                'cohorte_id': str(request.cohorte_id),
                'actividades': [a.nombre for a in actividades_a_procesar],
                'filas_sin_match': filas_sin_match,
            },
            filas_afectadas=len(calificaciones_data),
            materia_id=request.materia_id,
        )

        return CalificacionConfirmResponse(
            materia_id=request.materia_id,
            cohorte_id=request.cohorte_id,
            calificaciones_creadas=len(calificaciones_data),
            actividades_procesadas=[a.nombre for a in actividades_a_procesar],
        )

    async def import_reporte_finalizacion(
        self,
        materia_id: uuid.UUID,
        cohorte_id: uuid.UUID,
        archivo: UploadFile,
    ) -> ReporteFinalizacionPreviewResponse:
        content = await archivo.read()
        filename = (archivo.filename or '').lower()

        if filename.endswith('.xlsx'):
            headers, data_rows = _parse_calificaciones_xlsx(content)
        elif filename.endswith('.csv'):
            headers, data_rows = _parse_calificaciones_csv(content)
        else:
            raise HTTPException(
                status_code=400,
                detail='Formato no soportado. Usar archivos .xlsx o .csv',
            )

        actividades = _detectar_actividades(headers)

        version_activa = await self._version_repo.get_activa(materia_id, cohorte_id)
        if not version_activa:
            raise HTTPException(
                status_code=400,
                detail='No hay un padrón activo para esta materia y cohorte',
            )

        entradas = await self._entrada_repo.get_by_version(version_activa.id)
        entrada_map: dict[str, tuple[uuid.UUID, str, str]] = {}
        for e in entradas:
            key = f'{e.nombre.lower()}|{e.apellidos.lower()}|{e.email.lower()}'
            entrada_map[key] = (e.id, e.nombre, e.apellidos)

        col_nombre = _find_header_index(headers, 'nombre')
        if col_nombre < 0:
            col_nombre = _find_header_index(headers, 'Nombre')
        col_apellidos = _find_header_index(headers, 'apellidos')
        if col_apellidos < 0:
            col_apellidos = _find_header_index(headers, 'apellido')
        if col_apellidos < 0:
            col_apellidos = _find_header_index(headers, 'Apellidos')
        col_email = _find_header_index(headers, 'email')
        if col_email < 0:
            col_email = _find_header_index(headers, 'Email')

        textuales = [a for a in actividades if a.tipo == 'textual']
        entregas_sin_corregir = []

        for data_row in data_rows:
            nombre_val = _get_valor_columna(data_row, col_nombre) if col_nombre >= 0 else ''
            apellidos_val = _get_valor_columna(data_row, col_apellidos) if col_apellidos >= 0 else ''
            email_val = _get_valor_columna(data_row, col_email) if col_email >= 0 else ''
            key = f'{nombre_val.lower()}|{apellidos_val.lower()}|{email_val.lower()}'

            entrada_info = entrada_map.get(key)
            if entrada_info is None:
                continue
            entrada_id = entrada_info[0]

            for act in textuales:
                col_idx = _find_header_index_by_actividad(headers, act.nombre)
                if col_idx < 0:
                    continue
                raw_val = _get_valor_columna(data_row, col_idx)
                if not raw_val or raw_val.lower() not in ('entregado', 'completado', 'si', 'sí', 'presentó', 'realizado'):
                    continue

                existing = await self._calificacion_repo.get_by_actividad(
                    materia_id, cohorte_id, act.nombre,
                )
                existing_entrada_ids = {c.entrada_padron_id for c in existing}
                if entrada_id not in existing_entrada_ids:
                    entregas_sin_corregir.append(EntregaSinCorregir(
                        alumno_nombre=entrada_info[1],
                        alumno_apellidos=entrada_info[2],
                        actividad=act.nombre,
                        comision='',
                    ))

        return ReporteFinalizacionPreviewResponse(
            materia_id=materia_id,
            cohorte_id=cohorte_id,
            entregas_sin_corregir=entregas_sin_corregir,
            total=len(entregas_sin_corregir),
        )

    async def list_calificaciones(
        self,
        materia_id: uuid.UUID,
        cohorte_id: uuid.UUID | None = None,
        entrada_padron_id: uuid.UUID | None = None,
        umbral_pct: int = 60,
        valores_aprobatorios: set[str] | None = None,
    ) -> list[CalificacionResponse]:
        calificaciones = await self._calificacion_repo.get_by_materia(
            materia_id, cohorte_id, entrada_padron_id,
        )
        result = []
        for c in calificaciones:
            res = CalificacionResponse.model_validate(c)
            res.aprobado = _calcular_aprobado(
                float(c.nota_numerica) if c.nota_numerica is not None else None,
                c.nota_textual,
                umbral_pct,
                valores_aprobatorios,
            )
            result.append(res)
        return result

    async def get_umbral(
        self, materia_id: uuid.UUID, cohorte_id: uuid.UUID,
    ) -> tuple[int, set[str] | None]:
        umbral = await self._umbral_repo.get_by_materia_cohorte(materia_id, cohorte_id)
        if umbral:
            vals = set(umbral.valores_aprobatorios) if umbral.valores_aprobatorios else None
            return umbral.umbral_pct, vals
        return 60, None


def _find_header_index(headers: list[str], target: str) -> int:
    target_lower = target.lower()
    for i, h in enumerate(headers):
        if h.strip().lower() == target_lower:
            return i
    return -1


def _find_header_index_by_actividad(headers: list[str], actividad_nombre: str) -> int:
    nombre_lower = actividad_nombre.lower()
    for i, h in enumerate(headers):
        clean = h.strip()
        if clean.lower() == nombre_lower:
            return i
        if clean.endswith('(Real)') and clean[:-7].strip().lower() == nombre_lower:
            return i
    return -1
