import csv
import io
import uuid
from collections.abc import Sequence

from fastapi import HTTPException, UploadFile
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.audit_codes import PADRON_CARGAR
from app.models.calificacion import Calificacion
from app.models.user import User
from app.repositories.padron_repository import EntradaPadronRepository, VersionPadronRepository
from app.schemas.auth import UserContext
from app.schemas.padron import (
    ImportPreviewItem,
    ImportPreviewResponse,
    PadronVaciarResponse,
    VersionPadronResponse,
)
from app.services.audit_service import AuditService

COLUMNAS_ESPERADAS = {'nombre', 'apellidos', 'email', 'comision', 'regional'}


def _parse_xlsx(file_bytes: bytes) -> list[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail='El archivo está vacío')
    headers = [str(h).strip().lower() if h else '' for h in rows[0]]
    if not COLUMNAS_ESPERADAS.issubset(set(headers)):
        raise HTTPException(
            status_code=400,
            detail=f'Columnas requeridas: {", ".join(sorted(COLUMNAS_ESPERADAS))}. Encontradas: {", ".join(headers)}',
        )
    result = []
    for row in rows[1:]:
        if not any(row):
            continue
        item = {}
        for i, header in enumerate(headers):
            if header in COLUMNAS_ESPERADAS:
                item[header] = str(row[i]).strip() if i < len(row) and row[i] is not None else ''
        result.append(item)
    return result


def _parse_csv(file_bytes: bytes) -> list[dict]:
    text = file_bytes.decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(text))
    if reader.fieldnames is None:
        raise HTTPException(status_code=400, detail='No se pudieron leer las columnas del CSV')
    headers = [h.strip().lower() for h in reader.fieldnames]
    if not COLUMNAS_ESPERADAS.issubset(set(headers)):
        raise HTTPException(
            status_code=400,
            detail=f'Columnas requeridas: {", ".join(sorted(COLUMNAS_ESPERADAS))}. Encontradas: {", ".join(headers)}',
        )
    result = []
    for row in reader:
        item = {k.strip().lower(): (v or '').strip() for k, v in row.items() if k and k.strip().lower() in COLUMNAS_ESPERADAS}
        if any(item.get(h) for h in COLUMNAS_ESPERADAS):
            result.append(item)
    return result


class PadronService:
    def __init__(self, session: AsyncSession, tenant_id: uuid.UUID):
        self._session = session
        self._tenant_id = tenant_id
        self._version_repo = VersionPadronRepository(session, tenant_id)
        self._entrada_repo = EntradaPadronRepository(session, tenant_id)

    async def preview_import(
        self,
        materia_id: uuid.UUID,
        cohorte_id: uuid.UUID,
        archivo: UploadFile,
    ) -> ImportPreviewResponse:
        content = await archivo.read()
        filename = (archivo.filename or '').lower()

        if filename.endswith('.xlsx'):
            items = _parse_xlsx(content)
        elif filename.endswith('.csv'):
            items = _parse_csv(content)
        else:
            raise HTTPException(
                status_code=400,
                detail='Formato no soportado. Usar archivos .xlsx o .csv',
            )

        if not items:
            raise HTTPException(status_code=400, detail='El archivo no contiene datos válidos')

        return ImportPreviewResponse(
            materia_id=materia_id,
            cohorte_id=cohorte_id,
            total=len(items),
            items=[ImportPreviewItem(**item) for item in items],
        )

    async def confirm_import(
        self,
        materia_id: uuid.UUID,
        cohorte_id: uuid.UUID,
        items: list[ImportPreviewItem],
        current_user: UserContext,
        audit: AuditService,
    ) -> VersionPadronResponse:
        activa = await self._version_repo.get_activa(materia_id, cohorte_id)
        if activa:
            await self._version_repo.desactivar_version(activa.id)

        version = await self._version_repo.create(
            materia_id=materia_id,
            cohorte_id=cohorte_id,
            cargado_por=current_user.user_id,
            activa=True,
        )

        emails = list({i.email.strip().lower() for i in items if i.email})
        usuarios_map: dict[str, uuid.UUID] = {}
        if emails:
            result = await self._session.execute(
                select(User).where(
                    User.tenant_id == self._tenant_id,
                    User.deleted_at.is_(None),
                    User.email.in_(emails),
                ),
            )
            for u in result.scalars().all():
                usuarios_map[u.email.strip().lower()] = u.id

        entrada_data = []
        for item in items:
            email_normalized = item.email.strip().lower()
            entrada_data.append({
                'nombre': item.nombre,
                'apellidos': item.apellidos,
                'email': item.email,
                'comision': item.comision,
                'regional': item.regional,
                'usuario_id': usuarios_map.get(email_normalized),
            })

        await self._entrada_repo.bulk_create(version.id, entrada_data)

        await audit.log(
            accion=PADRON_CARGAR,
            detalle={
                'tipo': 'import',
                'materia_id': str(materia_id),
                'cohorte_id': str(cohorte_id),
                'version_id': str(version.id),
            },
            filas_afectadas=len(items),
            materia_id=materia_id,
        )

        return VersionPadronResponse.model_validate(version)

    async def vaciar_materia(
        self,
        materia_id: uuid.UUID,
        current_user: UserContext,
        audit: AuditService,
    ) -> PadronVaciarResponse:
        cohortes_result = await self._session.execute(
            select(self._version_repo._model.cohorte_id)
            .where(self._version_repo._model.materia_id == materia_id)
            .where(self._version_repo._model.tenant_id == self._tenant_id)
            .where(self._version_repo._model.deleted_at.is_(None))
            .where(self._version_repo._model.activa.is_(True))
            .distinct(),
        )
        cohorte_ids = [row[0] for row in cohortes_result.all()]

        total_entradas = 0
        version_desactivada = False

        for cohorte_id in cohorte_ids:
            activa = await self._version_repo.get_activa(materia_id, cohorte_id)
            if activa:
                count = await self._entrada_repo.soft_delete_by_version(activa.id)
                total_entradas += count
                await self._version_repo.desactivar_version(activa.id)
                version_desactivada = True

        # También eliminar calificaciones de esta materia+cohorte
        from sqlalchemy import delete as sa_delete
        await self._session.execute(
            sa_delete(Calificacion).where(
                Calificacion.materia_id == materia_id,
                Calificacion.cohorte_id.in_(cohorte_ids),
                Calificacion.tenant_id == self._tenant_id,
            ),
        )

        await audit.log(
            accion=PADRON_CARGAR,
            detalle={
                'tipo': 'vaciar',
                'materia_id': str(materia_id),
            },
            filas_afectadas=total_entradas,
            materia_id=materia_id,
        )

        return PadronVaciarResponse(
            materia_id=materia_id,
            entradas_afectadas=total_entradas,
            version_desactivada=version_desactivada,
        )

    async def list_versiones(self) -> list[VersionPadronResponse]:
        versiones = await self._version_repo.get_multi(limit=1000)
        return [VersionPadronResponse.model_validate(v) for v in versiones]
